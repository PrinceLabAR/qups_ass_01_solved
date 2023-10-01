import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define the functions to find the longest and shortest non-empty suggestions
def find_longest_non_empty_suggestion(suggestions):
    longest_suggestion = None

    for suggestion in suggestions:
        suggestion_text = suggestion.text.strip()

        if suggestion_text:
            if longest_suggestion is None or len(suggestion_text) > len(longest_suggestion):
                longest_suggestion = suggestion_text

    return longest_suggestion

def find_shortest_non_empty_suggestion(suggestions):
    shortest_suggestion = None

    for suggestion in suggestions:
        suggestion_text = suggestion.text.strip()

        if suggestion_text:
            if shortest_suggestion is None or len(suggestion_text) < len(shortest_suggestion):
                shortest_suggestion = suggestion_text

    return shortest_suggestion

# Step 01 - Get specific values from rows 3 to 12 in column 3 of the Excel file
excel_file = "Excel.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active
cell_values = [sheet.cell(row=row_num, column=3).value for row_num in range(3, 13)]

# Step 02 - Open Google and perform a search for each value
driver = webdriver.Chrome()

# Lists to store the longest and shortest suggestions for each value
longest_suggestions = []
shortest_suggestions = []

for value in cell_values:
    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(value)

    # Rest of your code for finding suggestions and printing them

    # Wait for suggestions and find the longest and shortest suggestions
    wait = WebDriverWait(driver, 5)
    wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "sbct")))
    suggestion_elements = driver.find_elements(By.CLASS_NAME, "sbct")

    longest_suggestion = find_longest_non_empty_suggestion(suggestion_elements)
    shortest_suggestion = find_shortest_non_empty_suggestion(suggestion_elements)

    # Append the results to the lists
    longest_suggestions.append(longest_suggestion)
    shortest_suggestions.append(shortest_suggestion)

# Step 03 - Write results back to the Excel file for rows 3 to 12
output_sheet = workbook["Monday"]
for row_num, longest, shortest in zip(range(3, 13), longest_suggestions, shortest_suggestions):
    output_sheet.cell(row=row_num, column=4, value=longest)
    output_sheet.cell(row=row_num, column=5, value=shortest)

# Step 04 - Save changes to the Excel file
workbook.save(excel_file)
workbook.close()

# Quit the WebDriver
driver.quit()
